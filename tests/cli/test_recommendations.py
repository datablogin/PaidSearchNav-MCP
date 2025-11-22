"""Tests for the recommendations CLI commands."""

import json
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from click.testing import CliRunner
from openpyxl import load_workbook

from paidsearchnav.cli.main import cli
from paidsearchnav.storage.models import AnalysisRecord, Audit, Customer


@pytest.fixture
def runner():
    """Create a Click test runner."""
    return CliRunner()


@pytest.fixture
def mock_audit():
    """Create a mock audit object."""
    customer = MagicMock(spec=Customer)
    customer.id = "customer-123"
    customer.name = "Test Customer"
    customer.google_ads_customer_id = "1234567890"

    audit = MagicMock(spec=Audit)
    audit.id = "audit-123"
    audit.name = "Test Audit"
    audit.status = "completed"
    audit.customer = customer
    audit.total_recommendations = 15
    audit.potential_savings = 2500.0

    return audit


@pytest.fixture
def mock_analysis_records():
    """Create mock analysis records with recommendations."""
    records = []

    # Search terms analyzer record
    search_record = MagicMock(spec=AnalysisRecord)
    search_record.analyzer_name = "search_terms"
    search_record.analysis_type = "search_terms"
    search_record.result_data = {
        "recommendations": [
            {
                "type": "ADD_NEGATIVE",
                "priority": "high",
                "title": "Add 'free' as negative keyword",
                "description": "Low-value search term with 0% conversion rate",
                "keyword": "free",
                "search_term": "free",
                "campaign": "Brand Campaign",
                "ad_group": "Brand Terms",
                "cost_savings": 500.0,
                "estimated_impact": "High",
            },
            {
                "type": "ADD_NEGATIVE",
                "priority": "medium",
                "title": "Add 'cheap' as negative keyword",
                "description": "Low-value search term",
                "keyword": "cheap",
                "search_term": "cheap",
                "campaign": "Brand Campaign",
                "cost_savings": 200.0,
            },
        ]
    }
    records.append(search_record)

    # Keyword match types analyzer record
    keyword_record = MagicMock(spec=AnalysisRecord)
    keyword_record.analyzer_name = "keyword_match_types"
    keyword_record.analysis_type = "keyword_match_types"
    keyword_record.result_data = {
        "recommendations": [
            {
                "type": "MATCH_TYPE_CHANGE",
                "priority": "high",
                "title": "Change 'shoes' from broad to phrase match",
                "description": "Improve targeting precision",
                "keyword": "shoes",
                "campaign": "Product Campaign",
                "ad_group": "Footwear",
                "current_match_type": "broad",
                "new_match_type": "phrase",
                "cost_savings": 800.0,
                "implementation_details": "Update in Google Ads Editor",
            },
            {
                "type": "PAUSE_KEYWORD",
                "priority": "high",
                "title": "Pause keyword 'discount shoes'",
                "description": "Poor performance, high cost",
                "keyword": "discount shoes",
                "campaign": "Product Campaign",
                "match_type": "broad",
                "cost_savings": 600.0,
            },
            {
                "type": "ADD_KEYWORD",
                "priority": "medium",
                "title": "Add 'running shoes near me'",
                "description": "High-intent local keyword",
                "keyword": "running shoes near me",
                "campaign": "Product Campaign",
                "ad_group": "Footwear",
                "match_type": "exact",
                "max_cpc": "2.50",
            },
        ]
    }
    records.append(keyword_record)

    # Bid optimization analyzer record
    bid_record = MagicMock(spec=AnalysisRecord)
    bid_record.analyzer_name = "bid_optimization"
    bid_record.analysis_type = "bid_optimization"
    bid_record.result_data = {
        "recommendations": [
            {
                "type": "BID_ADJUSTMENT",
                "priority": "medium",
                "title": "Decrease bid for 'shoes sale'",
                "description": "Reduce wasted spend",
                "keyword": "shoes sale",
                "current_bid": "3.00",
                "new_bid": "2.00",
                "adjustment": "-33%",
                "cost_savings": 400.0,
            }
        ]
    }
    records.append(bid_record)

    return records


def setup_mock_repo_and_session(
    mock_repo_class, mock_audit=None, mock_analysis_records=None
):
    """Helper to set up mock repository and session."""
    mock_repo = MagicMock()
    mock_repo_class.return_value = mock_repo

    # Create a mock session that properly handles async context manager
    mock_session = MagicMock()
    mock_async_session_context = MagicMock()
    mock_async_session_context.__aenter__ = AsyncMock(return_value=mock_session)
    mock_async_session_context.__aexit__ = AsyncMock(return_value=None)
    mock_repo.AsyncSessionLocal.return_value = mock_async_session_context

    # Mock database queries
    mock_session.get = AsyncMock(return_value=mock_audit)

    if mock_analysis_records is not None:
        # Mock the execute result chain
        mock_scalars = MagicMock()
        mock_scalars.all = MagicMock(return_value=mock_analysis_records)
        mock_execute_result = MagicMock()
        mock_execute_result.scalars.return_value = mock_scalars
        mock_session.execute = AsyncMock(return_value=mock_execute_result)

    return mock_repo, mock_session


class TestRecommendationsExport:
    """Test the recommendations export command."""

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_export_csv_success(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records, tmp_path
    ):
        """Test successful CSV export."""
        setup_mock_repo_and_session(mock_repo_class, mock_audit, mock_analysis_records)

        output_file = tmp_path / "recommendations.csv"
        result = runner.invoke(
            cli,
            [
                "recommendations",
                "export",
                "--audit-id",
                "audit-123",
                "--format",
                "csv",
                "--output",
                str(output_file),
            ],
        )

        assert result.exit_code == 0
        assert "Exported 6 recommendations" in result.output
        assert output_file.exists()

        # Check CSV content
        content = output_file.read_text()
        assert "Analyzer,Priority,Type,Title" in content
        assert "search_terms" in content
        assert "keyword_match_types" in content
        assert "free" in content
        assert "$500.00" in content

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_export_json_success(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records, tmp_path
    ):
        """Test successful JSON export."""
        setup_mock_repo_and_session(mock_repo_class, mock_audit, mock_analysis_records)

        output_file = tmp_path / "recommendations.json"
        result = runner.invoke(
            cli,
            [
                "recommendations",
                "export",
                "--audit-id",
                "audit-123",
                "--format",
                "json",
                "--output",
                str(output_file),
            ],
        )

        assert result.exit_code == 0
        assert output_file.exists()

        # Check JSON content
        data = json.loads(output_file.read_text())
        assert data["audit_id"] == "audit-123"
        assert data["customer"] == "Test Customer"
        assert data["total_recommendations"] == 6
        assert data["total_cost_savings"] == 2500.0
        assert len(data["recommendations"]) == 6

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_export_excel_success(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records, tmp_path
    ):
        """Test successful Excel export."""
        setup_mock_repo_and_session(mock_repo_class, mock_audit, mock_analysis_records)

        output_file = tmp_path / "recommendations.xlsx"
        result = runner.invoke(
            cli,
            [
                "recommendations",
                "export",
                "--audit-id",
                "audit-123",
                "--format",
                "excel",
                "--output",
                str(output_file),
            ],
        )

        assert result.exit_code == 0
        assert output_file.exists()

        # Check Excel content
        wb = load_workbook(output_file)
        assert "Summary" in wb.sheetnames
        assert "Recommendations" in wb.sheetnames
        assert "High Priority" in wb.sheetnames
        assert "Medium Priority" in wb.sheetnames

        # Check summary sheet
        summary_sheet = wb["Summary"]
        assert summary_sheet["A1"].value == "Audit Report"
        assert summary_sheet["B1"].value == "Test Audit"

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_export_google_ads_editor_success(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records, tmp_path
    ):
        """Test successful Google Ads Editor format export."""
        setup_mock_repo_and_session(mock_repo_class, mock_audit, mock_analysis_records)

        output_file = tmp_path / "editor_import.csv"
        result = runner.invoke(
            cli,
            [
                "recommendations",
                "export",
                "--audit-id",
                "audit-123",
                "--format",
                "google_ads_editor",
                "--output",
                str(output_file),
            ],
        )

        assert result.exit_code == 0
        assert output_file.exists()

        # Check Google Ads Editor format
        content = output_file.read_text()
        assert "Action,Campaign,Ad group,Keyword,Match type" in content
        assert "Add,Brand Campaign,Brand Terms,free,Negative exact" in content
        assert "Edit,Product Campaign,Footwear,shoes," in content
        assert "PaidSearchNav:" in content

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_export_with_priority_filter(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records, tmp_path
    ):
        """Test export with priority filtering."""
        setup_mock_repo_and_session(mock_repo_class, mock_audit, mock_analysis_records)

        output_file = tmp_path / "high_priority.csv"
        result = runner.invoke(
            cli,
            [
                "recommendations",
                "export",
                "--audit-id",
                "audit-123",
                "--format",
                "csv",
                "--output",
                str(output_file),
                "--priority",
                "high",
            ],
        )

        assert result.exit_code == 0
        assert "Exported 3 recommendations" in result.output

        # Check only high priority items are included
        content = output_file.read_text()
        assert "high" in content.lower()
        assert "medium" not in content.lower()

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_export_with_analyzer_filter(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records, tmp_path
    ):
        """Test export with analyzer filtering."""
        # Set up mocks with analyzer filter
        mock_repo = MagicMock()
        mock_repo_class.return_value = mock_repo

        mock_session = MagicMock()
        mock_async_session_context = MagicMock()
        mock_async_session_context.__aenter__ = AsyncMock(return_value=mock_session)
        mock_async_session_context.__aexit__ = AsyncMock(return_value=None)
        mock_repo.AsyncSessionLocal.return_value = mock_async_session_context

        mock_session.get = AsyncMock(return_value=mock_audit)

        # Return only search_terms records
        filtered_records = [
            r for r in mock_analysis_records if r.analyzer_name == "search_terms"
        ]
        mock_scalars = MagicMock()
        mock_scalars.all = MagicMock(return_value=filtered_records)
        mock_execute_result = MagicMock()
        mock_execute_result.scalars.return_value = mock_scalars
        mock_session.execute = AsyncMock(return_value=mock_execute_result)

        output_file = tmp_path / "search_terms_only.csv"
        result = runner.invoke(
            cli,
            [
                "recommendations",
                "export",
                "--audit-id",
                "audit-123",
                "--format",
                "csv",
                "--output",
                str(output_file),
                "--analyzer",
                "search_terms",
            ],
        )

        assert result.exit_code == 0
        assert "Exported 2 recommendations" in result.output

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_export_audit_not_found(self, mock_repo_class, runner, tmp_path):
        """Test export with non-existent audit."""
        setup_mock_repo_and_session(mock_repo_class, None, [])

        output_file = tmp_path / "recommendations.csv"
        result = runner.invoke(
            cli,
            [
                "recommendations",
                "export",
                "--audit-id",
                "non-existent",
                "--format",
                "csv",
                "--output",
                str(output_file),
            ],
        )

        assert result.exit_code == 0
        assert "Error: Audit non-existent not found" in result.output
        assert not output_file.exists()


class TestRecommendationsExportNegatives:
    """Test the export-negatives command."""

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_export_negatives_csv_success(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records, tmp_path
    ):
        """Test successful negative keyword export in CSV format."""
        setup_mock_repo_and_session(mock_repo_class, mock_audit, mock_analysis_records)

        output_file = tmp_path / "negatives.csv"
        result = runner.invoke(
            cli,
            [
                "recommendations",
                "export-negatives",
                "--audit-id",
                "audit-123",
                "--output",
                str(output_file),
            ],
        )

        assert result.exit_code == 0
        assert "Exported 2 negative keywords" in result.output
        assert output_file.exists()

        # Check CSV content
        content = output_file.read_text()
        assert "Keyword,Match Type,Campaign,Ad Group,Reason" in content
        assert "free,exact,Brand Campaign" in content
        assert "cheap,exact,Brand Campaign" in content

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_export_negatives_txt_success(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records, tmp_path
    ):
        """Test successful negative keyword export in TXT format."""
        setup_mock_repo_and_session(mock_repo_class, mock_audit, mock_analysis_records)

        output_file = tmp_path / "negatives.txt"
        result = runner.invoke(
            cli,
            [
                "recommendations",
                "export-negatives",
                "--audit-id",
                "audit-123",
                "--output",
                str(output_file),
                "--format",
                "txt",
            ],
        )

        assert result.exit_code == 0
        assert output_file.exists()

        # Check TXT content (formatted for Google Ads)
        content = output_file.read_text()
        assert "[free]" in content  # Exact match format
        assert "[cheap]" in content

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_export_negatives_no_negatives_found(
        self, mock_repo_class, runner, mock_audit, tmp_path
    ):
        """Test export when no negative keywords are found."""
        # Create records without negative keywords
        records = []
        record = MagicMock(spec=AnalysisRecord)
        record.analyzer_name = "other_analyzer"
        record.result_data = {"recommendations": [{"type": "OTHER_TYPE"}]}
        records.append(record)

        setup_mock_repo_and_session(mock_repo_class, mock_audit, records)

        output_file = tmp_path / "negatives.csv"
        result = runner.invoke(
            cli,
            [
                "recommendations",
                "export-negatives",
                "--audit-id",
                "audit-123",
                "--output",
                str(output_file),
            ],
        )

        assert result.exit_code == 0
        assert "No negative keyword recommendations found" in result.output
        assert not output_file.exists()


class TestRecommendationsPreview:
    """Test the preview command."""

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_preview_all_changes(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records
    ):
        """Test preview of all changes."""
        setup_mock_repo_and_session(mock_repo_class, mock_audit, mock_analysis_records)

        result = runner.invoke(
            cli,
            ["recommendations", "preview", "--audit-id", "audit-123"],
        )

        assert result.exit_code == 0
        assert "Change Preview: Test Audit" in result.output
        assert "Keyword Changes" in result.output
        assert "Negative Keywords to Add" in result.output
        assert "Bid Adjustments" in result.output
        assert "Total Changes to Apply: 6" in result.output

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_preview_keyword_changes_only(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records
    ):
        """Test preview of keyword changes only."""
        setup_mock_repo_and_session(mock_repo_class, mock_audit, mock_analysis_records)

        result = runner.invoke(
            cli,
            [
                "recommendations",
                "preview",
                "--audit-id",
                "audit-123",
                "--type",
                "keyword_changes",
            ],
        )

        assert result.exit_code == 0
        assert "Keyword Changes" in result.output
        assert "Match Type Change" in result.output
        assert "Pause Keyword" in result.output
        assert "Add Keyword" in result.output
        # Should not show other types
        assert "Negative Keywords to Add" not in result.output
        assert "Bid Adjustments" not in result.output

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_preview_with_limit(
        self, mock_repo_class, runner, mock_audit, mock_analysis_records
    ):
        """Test preview with limit."""
        setup_mock_repo_and_session(mock_repo_class, mock_audit, mock_analysis_records)

        result = runner.invoke(
            cli,
            ["recommendations", "preview", "--audit-id", "audit-123", "--limit", "2"],
        )

        assert result.exit_code == 0
        # Should show "and X more" messages
        assert "more keyword changes" in result.output


class TestRecommendationsValidate:
    """Test the validate command."""

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_validate_no_issues(self, mock_repo_class, runner, mock_audit):
        """Test validation with no issues."""
        # Create clean recommendations
        records = []
        record = MagicMock(spec=AnalysisRecord)
        record.analyzer_name = "test_analyzer"
        record.result_data = {
            "recommendations": [
                {
                    "type": "ADD_NEGATIVE",
                    "keyword": "unique1",
                    "priority": "low",
                    "cost_savings": 100,
                },
                {
                    "type": "ADD_NEGATIVE",
                    "keyword": "unique2",
                    "priority": "low",
                    "cost_savings": 100,
                },
            ]
        }
        records.append(record)

        setup_mock_repo_and_session(mock_repo_class, mock_audit, records)

        result = runner.invoke(
            cli,
            ["recommendations", "validate", "--audit-id", "audit-123"],
        )

        assert result.exit_code == 0
        assert "✓ Duplicate Negatives" in result.output
        assert "✓ Conflicting Actions" in result.output
        assert "All validation checks passed" in result.output

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_validate_with_duplicates(self, mock_repo_class, runner, mock_audit):
        """Test validation with duplicate negative keywords."""
        # Create recommendations with duplicates
        records = []
        for i in range(2):
            record = MagicMock(spec=AnalysisRecord)
            record.analyzer_name = f"analyzer_{i}"
            record.result_data = {
                "recommendations": [
                    {
                        "type": "ADD_NEGATIVE",
                        "keyword": "duplicate",  # Same keyword
                        "analyzer": f"analyzer_{i}",
                    }
                ]
            }
            records.append(record)

        setup_mock_repo_and_session(mock_repo_class, mock_audit, records)

        result = runner.invoke(
            cli,
            ["recommendations", "validate", "--audit-id", "audit-123"],
        )

        assert result.exit_code == 0
        assert "⚠ Duplicate Negatives" in result.output
        assert "WARNING" in result.output
        assert "duplicate negative keyword recommendations" in result.output

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_validate_with_conflicts(self, mock_repo_class, runner, mock_audit):
        """Test validation with conflicting keyword actions."""
        # Create conflicting recommendations
        records = []
        record = MagicMock(spec=AnalysisRecord)
        record.analyzer_name = "test_analyzer"
        record.result_data = {
            "recommendations": [
                {
                    "type": "PAUSE_KEYWORD",
                    "keyword": "conflict",
                    "analyzer": "analyzer1",
                },
                {
                    "type": "ADD_KEYWORD",
                    "keyword": "conflict",  # Same keyword, different action
                    "analyzer": "analyzer2",
                },
            ]
        }
        records.append(record)

        setup_mock_repo_and_session(mock_repo_class, mock_audit, records)

        result = runner.invoke(
            cli,
            ["recommendations", "validate", "--audit-id", "audit-123"],
        )

        assert result.exit_code == 0
        assert "✗ Conflicting Actions" in result.output
        assert "ERROR" in result.output
        assert "conflicting actions" in result.output
        assert "Errors found - review recommendations" in result.output

    @patch("paidsearchnav.cli.recommendations.AnalysisRepository")
    def test_validate_high_budget_impact(self, mock_repo_class, runner, mock_audit):
        """Test validation with high budget impact."""
        # Create high-impact recommendations
        records = []
        record = MagicMock(spec=AnalysisRecord)
        record.analyzer_name = "test_analyzer"
        record.result_data = {
            "recommendations": [
                {
                    "type": "PAUSE_KEYWORD",
                    "keyword": "expensive",
                    "priority": "high",
                    "cost_savings": 15000,  # High savings
                }
            ]
        }
        records.append(record)

        setup_mock_repo_and_session(mock_repo_class, mock_audit, records)

        result = runner.invoke(
            cli,
            ["recommendations", "validate", "--audit-id", "audit-123"],
        )

        assert result.exit_code == 0
        assert "⚠ Budget Impact" in result.output
        assert "WARNING" in result.output
        assert "$15,000.00" in result.output
        assert "significant budget impact" in result.output
