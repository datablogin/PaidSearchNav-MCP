"""Exporter for keyword match type analysis results."""

import csv
import io
import json
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from paidsearchnav.core.interfaces import Exporter
from paidsearchnav.core.models import KeywordMatchAnalysisResult


class KeywordMatchExporter(Exporter):
    """Export keyword match type analysis results to various formats."""

    def get_supported_formats(self) -> list[str]:
        """Get list of supported export formats."""
        return ["csv", "xlsx", "json"]

    def export(
        self,
        data: Any,
        filename: str,
        format: str = "csv",
        **kwargs: Any,
    ) -> bytes:
        """Export keyword match analysis results.

        Args:
            data: KeywordMatchAnalysisResult to export
            filename: Output filename
            format: Export format (csv, xlsx, json)
            **kwargs: Additional options:
                - include_details: Include detailed keyword lists (default: True)
                - include_duplicates: Include duplicate opportunities (default: True)

        Returns:
            Exported data as bytes

        Raises:
            ValueError: If unsupported format
        """
        if not isinstance(data, KeywordMatchAnalysisResult):
            raise ValueError("Data must be a KeywordMatchAnalysisResult")

        if format not in self.get_supported_formats():
            raise ValueError(f"Unsupported format: {format}")

        include_details = kwargs.get("include_details", True)
        include_duplicates = kwargs.get("include_duplicates", True)

        if format == "csv":
            return self._export_csv(data, include_details, include_duplicates)
        elif format == "xlsx":
            return self._export_xlsx(data, include_details, include_duplicates)
        else:  # json
            return self._export_json(data, include_details, include_duplicates)

    def _export_csv(
        self,
        data: KeywordMatchAnalysisResult,
        include_details: bool,
        include_duplicates: bool,
    ) -> bytes:
        """Export to CSV format."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Summary section
        writer.writerow(["Keyword Match Type Analysis Summary"])
        writer.writerow([])
        writer.writerow(["Analysis Date:", data.created_at.isoformat()])
        writer.writerow(
            ["Date Range:", f"{data.start_date.date()} to {data.end_date.date()}"]
        )
        writer.writerow(["Customer ID:", data.customer_id])
        writer.writerow(["Total Keywords Analyzed:", data.total_keywords])
        writer.writerow(
            ["Potential Monthly Savings:", f"${data.potential_savings:,.2f}"]
        )
        writer.writerow([])

        # Match type performance
        writer.writerow(["Match Type Performance"])
        writer.writerow(
            [
                "Match Type",
                "Count",
                "Impressions",
                "Clicks",
                "Cost",
                "Conversions",
                "CTR",
                "CPC",
                "CPA",
                "ROAS",
            ]
        )

        for match_type, stats in data.match_type_stats.items():
            writer.writerow(
                [
                    match_type,
                    stats["count"],
                    stats["impressions"],
                    stats["clicks"],
                    f"${stats['cost']:,.2f}",
                    f"{stats['conversions']:.1f}",
                    f"{stats['ctr']:.2f}%",
                    f"${stats['avg_cpc']:.2f}",
                    f"${stats['cpa']:.2f}" if stats["cpa"] > 0 else "N/A",
                    f"{stats['roas']:.2f}",
                ]
            )

        writer.writerow([])

        # Issues summary
        writer.writerow(["Issues Found"])
        writer.writerow(
            ["High-Cost Broad Keywords:", len(data.high_cost_broad_keywords)]
        )
        writer.writerow(["Low Quality Keywords:", len(data.low_quality_keywords)])
        writer.writerow(["Duplicate Opportunities:", len(data.duplicate_opportunities)])
        writer.writerow([])

        # Recommendations
        writer.writerow(["Recommendations"])
        for i, rec in enumerate(data.recommendations, 1):
            writer.writerow([f"{i}. [{rec.priority}] {rec.title}"])
            writer.writerow([f"   {rec.description}"])
        writer.writerow([])

        # Detailed sections
        if include_details:
            # High-cost broad keywords
            if data.high_cost_broad_keywords:
                writer.writerow(["High-Cost Broad Match Keywords"])
                writer.writerow(
                    [
                        "Keyword",
                        "Campaign",
                        "Ad Group",
                        "Impressions",
                        "Clicks",
                        "Cost",
                        "Conversions",
                        "CPA",
                        "ROAS",
                        "Quality Score",
                    ]
                )
                for keyword in data.high_cost_broad_keywords[:20]:  # Top 20
                    writer.writerow(
                        [
                            keyword.text,
                            keyword.campaign_name,
                            keyword.ad_group_name,
                            keyword.impressions,
                            keyword.clicks,
                            f"${keyword.cost:,.2f}",
                            f"{keyword.conversions:.1f}",
                            f"${keyword.cpa:.2f}" if keyword.cpa > 0 else "N/A",
                            f"{keyword.conversion_value / keyword.cost:.2f}"
                            if keyword.cost > 0
                            else "0.00",
                            keyword.quality_score or "N/A",
                        ]
                    )
                writer.writerow([])

            # Low quality keywords
            if data.low_quality_keywords:
                writer.writerow(["Low Quality Score Keywords"])
                writer.writerow(
                    [
                        "Keyword",
                        "Campaign",
                        "Ad Group",
                        "Match Type",
                        "Quality Score",
                        "Impressions",
                        "Cost",
                        "Conversions",
                    ]
                )
                for keyword in data.low_quality_keywords[:20]:  # Top 20
                    writer.writerow(
                        [
                            keyword.text,
                            keyword.campaign_name,
                            keyword.ad_group_name,
                            keyword.match_type,
                            keyword.quality_score,
                            keyword.impressions,
                            f"${keyword.cost:,.2f}",
                            f"{keyword.conversions:.1f}",
                        ]
                    )
                writer.writerow([])

        # Duplicate opportunities
        if include_duplicates and data.duplicate_opportunities:
            writer.writerow(["Duplicate Keyword Opportunities"])
            writer.writerow(
                [
                    "Keyword Text",
                    "Match Types Found",
                    "Recommended Match Type",
                    "Potential Savings",
                ]
            )
            for opp in data.duplicate_opportunities[:10]:  # Top 10
                writer.writerow(
                    [
                        opp["keyword_text"],
                        ", ".join(opp["match_types_found"]),
                        opp["recommended_match_type"],
                        f"${opp['potential_savings']:,.2f}",
                    ]
                )

        # Return CSV as bytes with BOM for Excel compatibility
        return output.getvalue().encode("utf-8-sig")

    def _export_xlsx(
        self,
        data: KeywordMatchAnalysisResult,
        include_details: bool,
        include_duplicates: bool,
    ) -> bytes:
        """Export to Excel format."""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Summary sheet
        summary_sheet = wb.create_sheet("Summary")
        self._create_summary_sheet(summary_sheet, data)

        # Match Type Performance sheet
        perf_sheet = wb.create_sheet("Match Type Performance")
        self._create_performance_sheet(perf_sheet, data)

        # Recommendations sheet
        rec_sheet = wb.create_sheet("Recommendations")
        self._create_recommendations_sheet(rec_sheet, data)

        if include_details:
            # High-cost broad keywords sheet
            if data.high_cost_broad_keywords:
                broad_sheet = wb.create_sheet("High-Cost Broad")
                self._create_broad_keywords_sheet(broad_sheet, data)

            # Low quality keywords sheet
            if data.low_quality_keywords:
                quality_sheet = wb.create_sheet("Low Quality")
                self._create_low_quality_sheet(quality_sheet, data)

        if include_duplicates and data.duplicate_opportunities:
            # Duplicate opportunities sheet
            dup_sheet = wb.create_sheet("Duplicates")
            self._create_duplicates_sheet(dup_sheet, data)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def _create_summary_sheet(self, sheet, data: KeywordMatchAnalysisResult):
        """Create summary sheet in Excel."""
        # Title
        sheet["A1"] = "Keyword Match Type Analysis Summary"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet.merge_cells("A1:B1")

        # Summary data
        summary_data = [
            ["Analysis Date", data.created_at.strftime("%Y-%m-%d %H:%M:%S")],
            ["Date Range", f"{data.start_date.date()} to {data.end_date.date()}"],
            ["Customer ID", data.customer_id],
            ["Total Keywords Analyzed", data.total_keywords],
            ["Potential Monthly Savings", f"${data.potential_savings:,.2f}"],
            ["", ""],
            ["Issues Found", ""],
            ["High-Cost Broad Keywords", len(data.high_cost_broad_keywords)],
            ["Low Quality Keywords", len(data.low_quality_keywords)],
            ["Duplicate Opportunities", len(data.duplicate_opportunities)],
        ]

        for row_idx, (label, value) in enumerate(summary_data, start=3):
            sheet[f"A{row_idx}"] = label
            sheet[f"B{row_idx}"] = value
            if label in ["Issues Found"]:
                sheet[f"A{row_idx}"].font = Font(bold=True)

        # Adjust column widths
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 30

    def _create_performance_sheet(self, sheet, data: KeywordMatchAnalysisResult):
        """Create match type performance sheet."""
        # Headers
        headers = [
            "Match Type",
            "Count",
            "Impressions",
            "Clicks",
            "Cost",
            "Conversions",
            "CTR",
            "Avg CPC",
            "CPA",
            "ROAS",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )

        # Data
        row_idx = 2
        for match_type, stats in data.match_type_stats.items():
            sheet.cell(row=row_idx, column=1, value=match_type)
            sheet.cell(row=row_idx, column=2, value=stats["count"])
            sheet.cell(row=row_idx, column=3, value=stats["impressions"])
            sheet.cell(row=row_idx, column=4, value=stats["clicks"])
            sheet.cell(row=row_idx, column=5, value=f"${stats['cost']:,.2f}")
            sheet.cell(row=row_idx, column=6, value=round(stats["conversions"], 1))
            sheet.cell(row=row_idx, column=7, value=f"{stats['ctr']:.2f}%")
            sheet.cell(row=row_idx, column=8, value=f"${stats['avg_cpc']:.2f}")
            sheet.cell(
                row=row_idx,
                column=9,
                value=f"${stats['cpa']:.2f}" if stats["cpa"] > 0 else "N/A",
            )
            sheet.cell(row=row_idx, column=10, value=round(stats["roas"], 2))
            row_idx += 1

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    def _create_recommendations_sheet(self, sheet, data: KeywordMatchAnalysisResult):
        """Create recommendations sheet."""
        # Headers
        headers = ["Priority", "Type", "Recommendation", "Description"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )

        # Data
        for row_idx, rec in enumerate(data.recommendations, start=2):
            sheet.cell(row=row_idx, column=1, value=rec.priority)
            sheet.cell(row=row_idx, column=2, value=rec.type)
            sheet.cell(row=row_idx, column=3, value=rec.title)
            sheet.cell(row=row_idx, column=4, value=rec.description)

            # Color code by priority
            if rec.priority == "HIGH":
                fill_color = "FFE0E0"  # Light red
            elif rec.priority == "MEDIUM":
                fill_color = "FFF0E0"  # Light orange
            else:
                fill_color = "E0F0E0"  # Light green

            for col in range(1, 5):
                sheet.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )

        # Adjust column widths
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 40
        sheet.column_dimensions["D"].width = 60

    def _create_broad_keywords_sheet(self, sheet, data: KeywordMatchAnalysisResult):
        """Create high-cost broad keywords sheet."""
        # Create DataFrame
        df_data = []
        for keyword in data.high_cost_broad_keywords[:50]:  # Top 50
            df_data.append(
                {
                    "Keyword": keyword.text,
                    "Campaign": keyword.campaign_name,
                    "Ad Group": keyword.ad_group_name,
                    "Impressions": keyword.impressions,
                    "Clicks": keyword.clicks,
                    "Cost": keyword.cost,
                    "Conversions": keyword.conversions,
                    "CPA": keyword.cpa if keyword.cpa > 0 else 0,
                    "ROAS": keyword.conversion_value / keyword.cost
                    if keyword.cost > 0
                    else 0,
                    "Quality Score": keyword.quality_score or 0,
                }
            )

        df = pd.DataFrame(df_data)

        # Write to sheet
        for r_idx, row in enumerate(df.values, start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Write headers
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = sheet.cell(row=1, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )

    def _create_low_quality_sheet(self, sheet, data: KeywordMatchAnalysisResult):
        """Create low quality keywords sheet."""
        # Headers
        headers = [
            "Keyword",
            "Campaign",
            "Ad Group",
            "Match Type",
            "Quality Score",
            "Impressions",
            "Cost",
            "Conversions",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )

        # Data
        for row_idx, keyword in enumerate(data.low_quality_keywords[:50], start=2):
            sheet.cell(row=row_idx, column=1, value=keyword.text)
            sheet.cell(row=row_idx, column=2, value=keyword.campaign_name)
            sheet.cell(row=row_idx, column=3, value=keyword.ad_group_name)
            sheet.cell(row=row_idx, column=4, value=keyword.match_type)
            sheet.cell(row=row_idx, column=5, value=keyword.quality_score)
            sheet.cell(row=row_idx, column=6, value=keyword.impressions)
            sheet.cell(row=row_idx, column=7, value=keyword.cost)
            sheet.cell(row=row_idx, column=8, value=keyword.conversions)

    def _create_duplicates_sheet(self, sheet, data: KeywordMatchAnalysisResult):
        """Create duplicate opportunities sheet."""
        # Headers
        headers = [
            "Keyword Text",
            "Match Types Found",
            "Recommended Match Type",
            "Potential Savings",
            "Details",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            )

        # Data
        for row_idx, opp in enumerate(data.duplicate_opportunities[:20], start=2):
            sheet.cell(row=row_idx, column=1, value=opp["keyword_text"])
            sheet.cell(row=row_idx, column=2, value=", ".join(opp["match_types_found"]))
            sheet.cell(row=row_idx, column=3, value=opp["recommended_match_type"])
            sheet.cell(row=row_idx, column=4, value=f"${opp['potential_savings']:,.2f}")

            # Build details string
            details = f"Found {len(opp['keywords'])} duplicate keywords"
            sheet.cell(row=row_idx, column=5, value=details)

    def _export_json(
        self,
        data: KeywordMatchAnalysisResult,
        include_details: bool,
        include_duplicates: bool,
    ) -> bytes:
        """Export to JSON format."""
        output = {
            "analysis_metadata": {
                "analysis_date": data.created_at.isoformat(),
                "date_range": {
                    "start": data.start_date.isoformat(),
                    "end": data.end_date.isoformat(),
                },
                "customer_id": data.customer_id,
                "analyzer": data.analyzer_name,
            },
            "summary": {
                "total_keywords": data.total_keywords,
                "potential_monthly_savings": data.potential_savings,
                "issues_found": {
                    "high_cost_broad": len(data.high_cost_broad_keywords),
                    "low_quality": len(data.low_quality_keywords),
                    "duplicate_opportunities": len(data.duplicate_opportunities),
                },
            },
            "match_type_performance": data.match_type_stats,
            "recommendations": [
                {
                    "priority": rec.priority,
                    "type": rec.type,
                    "title": rec.title,
                    "description": rec.description,
                }
                for rec in data.recommendations
            ],
        }

        if include_details:
            output["high_cost_broad_keywords"] = [
                {
                    "keyword": k.text,
                    "campaign": k.campaign_name,
                    "ad_group": k.ad_group_name,
                    "cost": k.cost,
                    "conversions": k.conversions,
                    "cpa": k.cpa if k.cpa > 0 else None,
                    "quality_score": k.quality_score,
                }
                for k in data.high_cost_broad_keywords[:20]
            ]

            output["low_quality_keywords"] = [
                {
                    "keyword": k.text,
                    "campaign": k.campaign_name,
                    "ad_group": k.ad_group_name,
                    "match_type": k.match_type,
                    "quality_score": k.quality_score,
                    "cost": k.cost,
                }
                for k in data.low_quality_keywords[:20]
            ]

        if include_duplicates:
            output["duplicate_opportunities"] = data.duplicate_opportunities[:10]

        return json.dumps(output, indent=2).encode("utf-8")
