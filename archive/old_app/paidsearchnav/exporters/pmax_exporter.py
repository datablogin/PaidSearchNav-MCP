"""Exporter for Performance Max analysis results."""

import csv
import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from paidsearchnav.core.interfaces import Exporter
from paidsearchnav.core.models.analysis import PerformanceMaxAnalysisResult


class PerformanceMaxExporter(Exporter):
    """Export Performance Max analysis results to various formats."""

    def get_supported_formats(self) -> list[str]:
        """Get list of supported export formats."""
        return ["csv", "xlsx", "json"]

    def export(
        self,
        data: Any,
        filename: str,
        format: str = "csv",
        **kwargs: Any,
    ) -> bytes:
        """Export Performance Max analysis results.

        Args:
            data: PerformanceMaxAnalysisResult to export
            filename: Output filename
            format: Export format (csv, xlsx, json)
            **kwargs: Additional options:
                - include_details: Include detailed search term lists (default: True)
                - include_overlaps: Include overlap analysis details (default: True)
                - include_recommendations: Include recommendations (default: True)

        Returns:
            Exported data as bytes

        Raises:
            ValueError: If unsupported format
        """
        if not isinstance(data, PerformanceMaxAnalysisResult):
            raise ValueError("Data must be a PerformanceMaxAnalysisResult")

        if format not in self.get_supported_formats():
            raise ValueError(f"Unsupported format: {format}")

        include_details = kwargs.get("include_details", True)
        include_overlaps = kwargs.get("include_overlaps", True)
        include_recommendations = kwargs.get("include_recommendations", True)

        if format == "csv":
            return self._export_csv(
                data, include_details, include_overlaps, include_recommendations
            )
        elif format == "xlsx":
            return self._export_xlsx(
                data, include_details, include_overlaps, include_recommendations
            )
        elif format == "json":
            return self._export_json(data)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _export_csv(
        self,
        result: PerformanceMaxAnalysisResult,
        include_details: bool,
        include_overlaps: bool,
        include_recommendations: bool,
    ) -> bytes:
        """Export to CSV format with multiple sections."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Summary section
        writer.writerow(["Performance Max Analysis Summary"])
        writer.writerow(
            ["Date Range", f"{result.start_date.date()} to {result.end_date.date()}"]
        )
        writer.writerow(["Customer ID", result.customer_id])
        writer.writerow(["Total PMax Campaigns", result.total_pmax_campaigns])
        writer.writerow(["Total Spend", f"${result.total_pmax_spend:.2f}"])
        writer.writerow(["Total Conversions", f"{result.total_pmax_conversions:.1f}"])
        writer.writerow(["Average ROAS", f"{result.avg_pmax_roas:.2f}"])
        writer.writerow(["Overlap Percentage", f"{result.overlap_percentage:.1f}%"])
        writer.writerow([])

        # Campaign Performance section
        if result.pmax_campaigns:
            writer.writerow(["Campaign Performance"])
            writer.writerow(
                [
                    "Campaign ID",
                    "Campaign Name",
                    "Spend",
                    "Conversions",
                    "Conversion Value",
                    "ROAS",
                    "Performance Status",
                ]
            )

            for campaign in result.pmax_campaigns:
                performance_status = (
                    "Good" if campaign.roas >= 1.5 else "Needs Optimization"
                )
                writer.writerow(
                    [
                        campaign.campaign_id,
                        campaign.name,
                        f"${campaign.cost:.2f}",
                        f"{campaign.conversions:.1f}",
                        f"${campaign.conversion_value:.2f}",
                        f"{campaign.roas:.2f}",
                        performance_status,
                    ]
                )
            writer.writerow([])

        # Search Terms Analysis section
        if include_details and result.search_term_analysis:
            search_analysis = result.search_term_analysis

            writer.writerow(["Search Terms Analysis"])
            writer.writerow(
                ["Total Search Terms", search_analysis.get("total_terms", 0)]
            )
            writer.writerow(
                ["High Volume Terms", len(search_analysis.get("high_volume_terms", []))]
            )
            writer.writerow(
                ["Irrelevant Terms", len(search_analysis.get("irrelevant_terms", []))]
            )
            writer.writerow(
                [
                    "Local Intent Terms",
                    len(search_analysis.get("local_intent_terms", [])),
                ]
            )
            writer.writerow([])

            # High volume terms details
            high_volume_terms = search_analysis.get("high_volume_terms", [])
            if high_volume_terms:
                writer.writerow(["Top High Volume Search Terms"])
                writer.writerow(
                    ["Search Term", "Impressions", "Clicks", "Cost", "Conversions"]
                )
                for term in high_volume_terms[:10]:  # Top 10
                    writer.writerow(
                        [
                            term.search_term,
                            term.metrics.impressions,
                            term.metrics.clicks,
                            f"${term.metrics.cost:.2f}",
                            f"{term.metrics.conversions:.1f}",
                        ]
                    )
                writer.writerow([])

            # Irrelevant terms details
            irrelevant_terms = search_analysis.get("irrelevant_terms", [])
            if irrelevant_terms:
                writer.writerow(
                    ["Irrelevant Search Terms (Negative Keyword Candidates)"]
                )
                writer.writerow(["Search Term", "Cost", "Clicks", "Conversions"])
                for term in irrelevant_terms[:20]:  # Top 20
                    writer.writerow(
                        [
                            term.search_term,
                            f"${term.metrics.cost:.2f}",
                            term.metrics.clicks,
                            f"{term.metrics.conversions:.1f}",
                        ]
                    )
                writer.writerow([])

        # Overlap Analysis section
        if include_overlaps and result.overlap_analysis:
            overlap_analysis = result.overlap_analysis

            writer.writerow(["Search/PMax Overlap Analysis"])
            writer.writerow(
                [
                    "Overlap Percentage",
                    f"{overlap_analysis.get('overlap_percentage', 0):.1f}%",
                ]
            )
            writer.writerow(
                [
                    "Total Overlapping Terms",
                    len(overlap_analysis.get("overlapping_terms", [])),
                ]
            )
            writer.writerow(
                [
                    "High Cost Overlaps",
                    len(overlap_analysis.get("high_cost_overlaps", [])),
                ]
            )
            writer.writerow([])

            # Overlapping terms details
            overlapping_terms = overlap_analysis.get("overlapping_terms", [])
            if overlapping_terms:
                writer.writerow(["Overlapping Search Terms"])
                writer.writerow(
                    ["Search Term", "PMax Cost", "Search Cost", "Total Cost"]
                )
                for term in overlapping_terms[:15]:  # Top 15
                    writer.writerow(
                        [
                            term["query"],
                            f"${term['pmax_cost']:.2f}",
                            f"${term['search_cost']:.2f}",
                            f"${term['total_cost']:.2f}",
                        ]
                    )
                writer.writerow([])

        # Findings section
        if result.findings:
            writer.writerow(["Key Findings"])
            writer.writerow(["Severity", "Type", "Title", "Description"])
            for finding in result.findings:
                writer.writerow(
                    [
                        finding.get("severity", ""),
                        finding.get("type", ""),
                        finding.get("title", ""),
                        finding.get("description", ""),
                    ]
                )
            writer.writerow([])

        # Recommendations section
        if include_recommendations and result.recommendations:
            writer.writerow(["Recommendations"])
            writer.writerow(
                ["Priority", "Type", "Title", "Description", "Estimated Impact"]
            )
            for rec in result.recommendations:
                writer.writerow(
                    [
                        rec.priority.value
                        if hasattr(rec.priority, "value")
                        else str(rec.priority),
                        rec.type.value if hasattr(rec.type, "value") else str(rec.type),
                        rec.title,
                        rec.description,
                        rec.estimated_impact or "",
                    ]
                )

        csv_content = output.getvalue()
        output.close()
        return csv_content.encode("utf-8")

    def _export_xlsx(
        self,
        result: PerformanceMaxAnalysisResult,
        include_details: bool,
        include_overlaps: bool,
        include_recommendations: bool,
    ) -> bytes:
        """Export to Excel format with multiple worksheets."""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Define styles
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        subheader_fill = PatternFill(
            start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
        )
        warning_fill = PatternFill(
            start_color="FFE699", end_color="FFE699", fill_type="solid"
        )
        critical_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

        # Summary worksheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Performance Max Analysis Summary"])
        ws_summary["A1"].fill = header_fill
        ws_summary["A1"].font = header_font

        summary_data = [
            [
                "Analysis Period",
                f"{result.start_date.date()} to {result.end_date.date()}",
            ],
            ["Customer ID", result.customer_id],
            ["Total PMax Campaigns", result.total_pmax_campaigns],
            ["Total Spend", f"${result.total_pmax_spend:.2f}"],
            ["Total Conversions", f"{result.total_pmax_conversions:.1f}"],
            ["Average ROAS", f"{result.avg_pmax_roas:.2f}"],
            ["Search Term Overlap", f"{result.overlap_percentage:.1f}%"],
        ]

        for row_data in summary_data:
            ws_summary.append(row_data)

        # Campaign Performance worksheet
        if result.pmax_campaigns:
            ws_campaigns = wb.create_sheet("Campaign Performance")
            headers = [
                "Campaign ID",
                "Campaign Name",
                "Spend",
                "Conversions",
                "Conversion Value",
                "ROAS",
                "Performance Status",
            ]
            ws_campaigns.append(headers)

            # Style headers
            for col, _header in enumerate(headers, 1):
                cell = ws_campaigns.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            for campaign in result.pmax_campaigns:
                performance_status = (
                    "Good" if campaign.roas >= 1.5 else "Needs Optimization"
                )
                row_data = [
                    campaign.campaign_id,
                    campaign.name,
                    campaign.cost,
                    campaign.conversions,
                    campaign.conversion_value,
                    campaign.roas,
                    performance_status,
                ]
                ws_campaigns.append(row_data)

                # Highlight poor performance
                if campaign.roas < 1.5:
                    for col in range(1, len(headers) + 1):
                        ws_campaigns.cell(
                            row=ws_campaigns.max_row, column=col
                        ).fill = warning_fill

        # Search Terms worksheet
        if include_details and result.search_term_analysis:
            search_analysis = result.search_term_analysis

            # High volume terms
            high_volume_terms = search_analysis.get("high_volume_terms", [])
            if high_volume_terms:
                ws_search = wb.create_sheet("High Volume Terms")
                headers = [
                    "Search Term",
                    "Impressions",
                    "Clicks",
                    "Cost",
                    "Conversions",
                    "CTR",
                    "CPA",
                ]
                ws_search.append(headers)

                for col, _header in enumerate(headers, 1):
                    cell = ws_search.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font

                for term in high_volume_terms[:50]:  # Top 50
                    row_data = [
                        term.search_term,
                        term.metrics.impressions,
                        term.metrics.clicks,
                        term.metrics.cost,
                        term.metrics.conversions,
                        f"{term.metrics.ctr:.2f}%",
                        f"${term.metrics.cpa:.2f}"
                        if term.metrics.conversions > 0
                        else "N/A",
                    ]
                    ws_search.append(row_data)

            # Irrelevant terms
            irrelevant_terms = search_analysis.get("irrelevant_terms", [])
            if irrelevant_terms:
                ws_negatives = wb.create_sheet("Negative Candidates")
                headers = [
                    "Search Term",
                    "Cost",
                    "Clicks",
                    "Conversions",
                    "Recommendation",
                ]
                ws_negatives.append(headers)

                for col, _header in enumerate(headers, 1):
                    cell = ws_negatives.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font

                for term in irrelevant_terms:
                    row_data = [
                        term.search_term,
                        term.metrics.cost,
                        term.metrics.clicks,
                        term.metrics.conversions,
                        "Add as negative keyword",
                    ]
                    ws_negatives.append(row_data)

                    # Highlight high-cost irrelevant terms
                    if term.metrics.cost > 50:
                        for col in range(1, len(headers) + 1):
                            ws_negatives.cell(
                                row=ws_negatives.max_row, column=col
                            ).fill = critical_fill

        # Overlap Analysis worksheet
        if include_overlaps and result.overlap_analysis:
            overlap_analysis = result.overlap_analysis
            overlapping_terms = overlap_analysis.get("overlapping_terms", [])

            if overlapping_terms:
                ws_overlap = wb.create_sheet("Search Overlap")
                headers = [
                    "Search Term",
                    "PMax Cost",
                    "Search Cost",
                    "Total Cost",
                    "Recommendation",
                ]
                ws_overlap.append(headers)

                for col, _header in enumerate(headers, 1):
                    cell = ws_overlap.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font

                for term in overlapping_terms:
                    recommendation = (
                        "Review keyword strategy"
                        if term["total_cost"] > 50
                        else "Monitor"
                    )
                    row_data = [
                        term["query"],
                        term["pmax_cost"],
                        term["search_cost"],
                        term["total_cost"],
                        recommendation,
                    ]
                    ws_overlap.append(row_data)

                    # Highlight high-cost overlaps
                    if term["total_cost"] > 100:
                        for col in range(1, len(headers) + 1):
                            ws_overlap.cell(
                                row=ws_overlap.max_row, column=col
                            ).fill = critical_fill

        # Recommendations worksheet
        if include_recommendations and result.recommendations:
            ws_recs = wb.create_sheet("Recommendations")
            headers = ["Priority", "Type", "Title", "Description", "Estimated Impact"]
            ws_recs.append(headers)

            for col, _header in enumerate(headers, 1):
                cell = ws_recs.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            for rec in result.recommendations:
                priority_val = (
                    rec.priority.value
                    if hasattr(rec.priority, "value")
                    else str(rec.priority)
                )
                type_val = (
                    rec.type.value if hasattr(rec.type, "value") else str(rec.type)
                )

                row_data = [
                    priority_val,
                    type_val,
                    rec.title,
                    rec.description,
                    rec.estimated_impact or "",
                ]
                ws_recs.append(row_data)

                # Color code by priority
                if priority_val == "CRITICAL":
                    fill = critical_fill
                elif priority_val == "HIGH":
                    fill = warning_fill
                else:
                    fill = subheader_fill

                for col in range(1, len(headers) + 1):
                    ws_recs.cell(row=ws_recs.max_row, column=col).fill = fill

        # Auto-adjust column widths
        for worksheet in wb.worksheets:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _export_json(self, result: PerformanceMaxAnalysisResult) -> bytes:
        """Export to JSON format."""
        # Convert to dictionary for JSON serialization
        data = {
            "analysis_summary": {
                "analyzer_name": result.analyzer_name,
                "customer_id": result.customer_id,
                "analysis_date": result.created_at.isoformat(),
                "date_range": {
                    "start_date": result.start_date.isoformat(),
                    "end_date": result.end_date.isoformat(),
                },
                "total_pmax_campaigns": result.total_pmax_campaigns,
                "total_pmax_spend": result.total_pmax_spend,
                "total_pmax_conversions": result.total_pmax_conversions,
                "avg_pmax_roas": result.avg_pmax_roas,
                "overlap_percentage": result.overlap_percentage,
            },
            "campaign_performance": [
                {
                    "campaign_id": campaign.campaign_id,
                    "campaign_name": campaign.name,
                    "spend": campaign.cost,
                    "conversions": campaign.conversions,
                    "conversion_value": campaign.conversion_value,
                    "roas": campaign.roas,
                    "performance_status": "Good"
                    if campaign.roas >= 1.5
                    else "Needs Optimization",
                }
                for campaign in result.pmax_campaigns
            ],
            "search_term_analysis": result.search_term_analysis,
            "overlap_analysis": result.overlap_analysis,
            "findings": result.findings,
            "recommendations": [
                {
                    "priority": rec.priority.value
                    if hasattr(rec.priority, "value")
                    else str(rec.priority),
                    "type": rec.type.value
                    if hasattr(rec.type, "value")
                    else str(rec.type),
                    "title": rec.title,
                    "description": rec.description,
                    "estimated_impact": rec.estimated_impact,
                    "campaign_id": rec.campaign_id,
                    "action_data": rec.action_data,
                }
                for rec in result.recommendations
            ],
            "metrics": result.metrics,
            "summary": result.summary,
        }

        json_content = json.dumps(data, indent=2, default=str)
        return json_content.encode("utf-8")
